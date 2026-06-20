# -*- coding: utf-8 -*-
"""
APP Seal - Extractor final dinámico de cortes programados SEAL Arequipa

Mejoras de esta versión:
- NO usa scan-from ni scan-to.
- Detecta dinámicamente los IDs desde el calendario renderizado.
- Abre cada ID real de DispForm.aspx y extrae el contenido completo.
- Corrige lectura de campos cuando SharePoint los muestra como:
    Título   ...
    Ubicación   ...
    Hora de inicio   ...
- Extrae descripción completa, zonas afectadas completas y SED/subestaciones.
- Guarda CSV, JSON, TXT, SQLite y XLSX con tabla de Excel.
- Permite fecha manual temporal. Si está vacía, usa fecha actual.
- Maneja una ventana máxima de 2 meses:
    30 días atrás + 30 días adelante.
- Limpia del historial local cortes anteriores a la ventana activa.

Instalación:
    pip install requests beautifulsoup4 playwright openpyxl
    python -m playwright install chromium

Uso normal:
    python seal_cortes_final_v2.py

Uso temporal para probar desde el 1 de junio:
    python seal_cortes_final_v2.py --fecha-consulta 2026-06-01

Uso desde el inicio del mes de la fecha actual:
    python seal_cortes_final_v2.py --desde-inicio-mes

Uso futuro real, solo desde hoy hacia adelante:
    python seal_cortes_final_v2.py --dias-atras 0 --dias-adelante 30

Depuración viendo navegador:
    python seal_cortes_final_v2.py --ver-navegador
"""

import argparse
import csv
import hashlib
import html
import json
import re
import sqlite3
import sys
import time
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import quote

import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# ============================================================
# CONFIGURACIÓN EDITABLE
# ============================================================

BASE_URL = "https://www.seal.com.pe"
SITE_URL = f"{BASE_URL}/clientes"
CORTES_PAGE = f"{SITE_URL}/SitePages/Cortes.aspx"
CALENDAR_PAGE = f"{SITE_URL}/Lists/Calendario/calendar.aspx"

DISPFORM_URL = (
    f"{SITE_URL}/Lists/Calendario/DispForm.aspx"
    f"?ID={{id}}&Source={quote(CORTES_PAGE, safe='')}"
)

OUTPUT_DIR_DEFAULT = r"C:\Users\Majuel-PC\Downloads\PY\APP_Seal\Resultados"

# Fecha manual temporal.
# Si quieres dejarlo fijo para pruebas:
# FECHA_MANUAL_DEFAULT = "2026-06-01"
# Si está vacío, usa la fecha actual.
FECHA_MANUAL_DEFAULT = ""

DIAS_ATRAS_DEFAULT = 7
DIAS_ADELANTE_DEFAULT = 14

MAX_DIAS_ATRAS = 14
MAX_DIAS_ADELANTE = 30

REQUEST_TIMEOUT = 30
PLAYWRIGHT_TIMEOUT = 90000
PLAYWRIGHT_WAIT_MS = 3500
SLEEP_BETWEEN_EVENTS = 0.05

SAVE_EVENT_HTML_DEFAULT = True

HEADERS_HTML = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) APP-Seal-Final-v2/4.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-PE,es;q=0.9,en;q=0.8",
    "Connection": "keep-alive",
}

CSV_COLUMNS = [
    "id",
    "fecha",
    "hora_inicio",
    "hora_fin",
    "distrito",
    "zona_afectada",
    "motivo",
    "ubicacion",
    "titulo",
    "subestaciones",
    "descripcion",
    "categoria",
    "todo_el_dia",
    "periodicidad",
    "datos_adjuntos",
    "tipo_contenido",
    "creado",
    "modificado",
    "creado_por",
    "modificado_por",
    "inicio_raw",
    "fin_raw",
    "inicio_iso",
    "fin_iso",
    "url",
    "fuente_extraccion",
    "hash",
]

XLSX_COLUMNS = [
    ("ID", "id"),
    ("Fecha", "fecha"),
    ("Hora inicio", "hora_inicio"),
    ("Hora fin", "hora_fin"),
    ("Distrito", "distrito"),
    ("Zona afectada", "zona_afectada"),
    ("Motivo", "motivo"),
    ("Ubicación", "ubicacion"),
    ("Título", "titulo"),
    ("Subestaciones / SED", "subestaciones"),
    ("Descripción completa", "descripcion"),
    ("Categoría", "categoria"),
    ("Creado", "creado"),
    ("Modificado", "modificado"),
    ("URL", "url"),
    ("Hash", "hash"),
]


# ============================================================
# UTILIDADES BÁSICAS
# ============================================================

def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def clean_text(value) -> str:
    if value is None:
        return ""

    value = str(value)
    value = html.unescape(value)
    value = value.replace("\xa0", " ")

    try:
        value = BeautifulSoup(value, "html.parser").get_text(" ", strip=True)
    except Exception:
        pass

    value = re.sub(r"\s+", " ", value)
    return value.strip()


def clean_multiline(value) -> str:
    if value is None:
        return ""

    value = str(value)
    value = html.unescape(value)
    value = value.replace("\xa0", " ")

    try:
        value = BeautifulSoup(value, "html.parser").get_text("\n", strip=True)
    except Exception:
        pass

    lines = []
    for line in value.splitlines():
        line = re.sub(r"\s+", " ", line).strip()
        if line:
            lines.append(line)

    return "\n".join(lines).strip()


def normalize_for_search(value: str) -> str:
    value = clean_text(value).lower()
    value = unicodedata.normalize("NFD", value)
    value = "".join(
        ch for ch in value
        if unicodedata.category(ch) != "Mn"
    )
    return value


def content_hash(*parts: str) -> str:
    raw = "|".join(clean_text(p) for p in parts)
    return hashlib.sha256(raw.encode("utf-8", errors="ignore")).hexdigest()


def shorten_console(value: str, width: int) -> str:
    value = clean_text(value)
    if len(value) <= width:
        return value
    return value[: max(0, width - 1)] + "…"


def choose_longer(*values: str) -> str:
    cleaned = [clean_multiline(v) for v in values if clean_multiline(v)]
    if not cleaned:
        return ""
    return max(cleaned, key=len)


def create_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS_HTML)

    retry = Retry(
        total=4,
        connect=4,
        read=4,
        status=4,
        backoff_factor=0.7,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        raise_on_status=False,
    )

    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)

    return session


def parse_yyyy_mm_dd(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def resolver_fecha_base(fecha_cli: Optional[str], usar_inicio_mes: bool) -> date:
    """
    Prioridad:
    1. --fecha-consulta
    2. FECHA_MANUAL_DEFAULT
    3. fecha actual
    """

    fecha_txt = clean_text(fecha_cli) or clean_text(FECHA_MANUAL_DEFAULT)

    if fecha_txt:
        fecha_base = parse_yyyy_mm_dd(fecha_txt)
    else:
        fecha_base = date.today()

    if usar_inicio_mes:
        fecha_base = fecha_base.replace(day=1)

    return fecha_base


def resolver_rango_consulta(
    fecha_base: date,
    dias_atras: int,
    dias_adelante: int,
) -> Tuple[date, date, int, int]:
    if dias_atras < 0:
        dias_atras = 0
    if dias_adelante < 0:
        dias_adelante = 0

    if dias_atras > MAX_DIAS_ATRAS:
        print(
            f"[AVISO] Solicitaste {dias_atras} días hacia atrás, "
            f"pero el máximo es {MAX_DIAS_ATRAS}. Se usará {MAX_DIAS_ATRAS}."
        )
        dias_atras = MAX_DIAS_ATRAS

    if dias_adelante > MAX_DIAS_ADELANTE:
        print(
            f"[AVISO] Solicitaste {dias_adelante} días hacia adelante, "
            f"pero el máximo es {MAX_DIAS_ADELANTE}. Se usará {MAX_DIAS_ADELANTE}."
        )
        dias_adelante = MAX_DIAS_ADELANTE

    fecha_inicio = fecha_base - timedelta(days=dias_atras)
    fecha_fin = fecha_base + timedelta(days=dias_adelante)

    return fecha_inicio, fecha_fin, dias_atras, dias_adelante


def month_starts_between(fecha_inicio: date, fecha_fin: date) -> List[date]:
    current = fecha_inicio.replace(day=1)
    end_month = fecha_fin.replace(day=1)
    months = []

    while current <= end_month:
        months.append(current)
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

    return months


def candidate_calendar_urls(month_start: date) -> List[str]:
    d = month_start.isoformat()
    return [
        f"{CORTES_PAGE}?CalendarDate={d}&CalendarPeriod=month",
        f"{CALENDAR_PAGE}?CalendarDate={d}&CalendarPeriod=month",
        f"{CORTES_PAGE}?CalendarDate={d}",
        f"{CALENDAR_PAGE}?CalendarDate={d}",
        CORTES_PAGE,
    ]


# ============================================================
# EXTRACCIÓN DINÁMICA DE IDS
# ============================================================

def extract_ids_from_text(text: str) -> Set[int]:
    ids: Set[int] = set()

    patterns = [
        r"DispForm\.aspx\?[^\"'<>\\]*?\bID=(\d+)",
        r"DispForm\.aspx\?[^\"'<>\\]*?\bid=(\d+)",
        r"Lists/Calendario/DispForm\.aspx\?ID=(\d+)",
        r"/clientes/Lists/Calendario/DispForm\.aspx\?ID=(\d+)",
    ]

    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            try:
                ids.add(int(match.group(1)))
            except Exception:
                pass

    return ids


def extract_id_from_href(href: str) -> Optional[int]:
    match = re.search(r"[?&]ID=(\d+)", href, flags=re.IGNORECASE)
    if match:
        return int(match.group(1))
    return None


def discover_ids_static_requests(
    session: requests.Session,
    fecha_inicio: date,
    fecha_fin: date,
    log,
) -> List[int]:
    """
    Fallback ligero. Puede encontrar cero IDs si el calendario depende de JS.
    """

    ids: Set[int] = set()

    for month_start in month_starts_between(fecha_inicio, fecha_fin):
        for url in candidate_calendar_urls(month_start):
            try:
                print(f"[INFO] Intento ligero sin navegador: {url}")
                response = session.get(url, headers=HEADERS_HTML, timeout=REQUEST_TIMEOUT)
                log.write(f"[STATIC] {url} -> HTTP {response.status_code}\n")

                found = set()
                if response.status_code == 200:
                    found = extract_ids_from_text(response.text)
                    ids.update(found)
                    log.write(f"[STATIC IDS] {sorted(found)}\n")

                if found:
                    break

            except Exception as exc:
                log.write(f"[STATIC ERROR] {url}: {repr(exc)}\n")

    return sorted(ids)


def collect_ids_from_playwright_page(page, log) -> Set[int]:
    ids: Set[int] = set()

    try:
        content = page.content()
        ids.update(extract_ids_from_text(content))
    except Exception as exc:
        log.write(f"[PLAYWRIGHT CONTENT ERROR] {repr(exc)}\n")

    try:
        hrefs = page.eval_on_selector_all(
            "a",
            "els => els.map(e => e.href || e.getAttribute('href') || '').filter(Boolean)"
        )

        for href in hrefs:
            item_id = extract_id_from_href(href)
            if item_id:
                ids.add(item_id)

    except Exception as exc:
        log.write(f"[PLAYWRIGHT HREF ERROR] {repr(exc)}\n")

    return ids


def click_more_links_and_collect(page, log) -> Set[int]:
    """
    En SharePoint hay enlaces tipo "1 elemento más".
    Intenta abrirlos para revelar todos los eventos ocultos del día.
    """

    ids: Set[int] = set()

    try:
        count = page.locator("a").count()
    except Exception:
        return ids

    max_clicks = min(count, 120)
    clicked = 0

    for _ in range(max_clicks):
        try:
            candidates = page.locator(
                "a",
                has_text=re.compile(r"elemento[s]?\s+m[aá]s", re.IGNORECASE)
            )

            candidate_count = candidates.count()
            if candidate_count <= clicked:
                break

            before_url = page.url
            candidates.nth(clicked).click(timeout=4000)
            page.wait_for_timeout(1400)

            ids.update(collect_ids_from_playwright_page(page, log))

            after_url = page.url
            if after_url != before_url:
                try:
                    page.go_back(wait_until="networkidle", timeout=30000)
                    page.wait_for_timeout(1000)
                except Exception:
                    try:
                        page.goto(before_url, wait_until="networkidle", timeout=30000)
                    except Exception:
                        pass
            else:
                try:
                    page.keyboard.press("Escape")
                    page.wait_for_timeout(300)
                except Exception:
                    pass

            clicked += 1

        except Exception as exc:
            log.write(f"[MORE CLICK ERROR] {repr(exc)}\n")
            clicked += 1

    return ids


def sync_playwright_cookies_to_requests(context, session: requests.Session) -> None:
    try:
        for c in context.cookies():
            name = c.get("name")
            value = c.get("value")
            domain = c.get("domain", "").lstrip(".")
            path = c.get("path", "/")
            if name and value:
                session.cookies.set(name, value, domain=domain, path=path)
    except Exception:
        pass


def create_playwright_context(headless: bool):
    from playwright.sync_api import sync_playwright

    p = sync_playwright().start()
    browser = p.chromium.launch(headless=headless)
    context = browser.new_context(
        user_agent=HEADERS_HTML["User-Agent"],
        locale="es-PE",
        viewport={"width": 1600, "height": 1000},
    )
    page = context.new_page()
    return p, browser, context, page


def discover_ids_rendered_with_playwright(
    page,
    session: requests.Session,
    fecha_inicio: date,
    fecha_fin: date,
    log,
) -> List[int]:
    ids: Set[int] = set()
    months = month_starts_between(fecha_inicio, fecha_fin)

    for month_start in months:
        urls = candidate_calendar_urls(month_start)
        log.write(f"\n[MONTH] {month_start.strftime('%Y-%m')}\n")
        month_ids_before = len(ids)

        for url in urls:
            try:
                print(f"[INFO] Renderizando calendario: {url}")
                log.write(f"[PLAYWRIGHT OPEN] {url}\n")

                page.goto(url, wait_until="networkidle", timeout=PLAYWRIGHT_TIMEOUT)
                page.wait_for_timeout(PLAYWRIGHT_WAIT_MS)

                found = collect_ids_from_playwright_page(page, log)
                ids.update(found)
                log.write(f"[PLAYWRIGHT IDS DIRECTOS] {sorted(found)}\n")

                more_found = click_more_links_and_collect(page, log)
                ids.update(more_found)

                if more_found:
                    log.write(f"[PLAYWRIGHT IDS MORE] {sorted(more_found)}\n")

                # Si esta URL ya devolvió IDs del mes, no probamos más variantes.
                if len(ids) > month_ids_before:
                    break

            except Exception as exc:
                log.write(f"[PLAYWRIGHT OPEN ERROR] {url}: {repr(exc)}\n")

    return sorted(ids)


# ============================================================
# PARSEO COMPLETO DEL DISPFORM
# ============================================================

KNOWN_LABELS = [
    "Título",
    "Ubicación",
    "Hora de inicio",
    "Hora de finalización",
    "Descripción",
    "Categoría",
    "Todo el día",
    "Periodicidad",
    "Datos adjuntos",
    "Tipo de contenido",
    "Creado",
    "Modificado",
    "Creado por",
    "Modificado por",
]

STOP_LABELS_FOR_DESCRIPTION = [
    "Categoría",
    "Todo el día",
    "Periodicidad",
    "Datos adjuntos",
    "Tipo de contenido",
    "Creado",
    "Modificado",
    "Creado el",
    "Última modificación",
    "Ultima modificacion",
    "Utilice esta página",
    "Utilice esta pagina",
    "Nombre",
]


def label_regex(label: str) -> re.Pattern:
    escaped = re.escape(label)
    return re.compile(rf"^\s*{escaped}\s*:?\s*(.*)$", re.IGNORECASE)


def is_known_label_line(value: str) -> bool:
    value_norm = normalize_for_search(value).rstrip(":")
    labels = KNOWN_LABELS + STOP_LABELS_FOR_DESCRIPTION

    for label in labels:
        label_norm = normalize_for_search(label).rstrip(":")
        if value_norm == label_norm:
            return True
        if value_norm.startswith(label_norm + " "):
            return True
        if value_norm.startswith(label_norm + ":"):
            return True

    return False


def get_page_text(soup: BeautifulSoup) -> str:
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()

    text = soup.get_text("\n", strip=True)
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)

    return text.strip()


def normalize_page_text_for_fields(text: str) -> str:
    """
    Convierte líneas que vienen como:
        Título   REFORZAMIENTO DE REDES.
    en algo más detectable, pero conserva descripción completa.
    """

    lines = []
    for line in text.splitlines():
        line = re.sub(r"[ \t]+", " ", line).strip()
        if line:
            lines.append(line)
    return "\n".join(lines)


def find_value_after_label(text: str, label: str, multiline: bool = False) -> str:
    """
    Soporta estos formatos:
    1) Título
       REFORZAMIENTO...
    2) Título   REFORZAMIENTO...
    3) Tipo de contenido: Evento
    """

    text = normalize_page_text_for_fields(text)
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    label_norm = normalize_for_search(label).rstrip(":")

    for i, line in enumerate(lines):
        line_clean = clean_text(line)
        line_norm = normalize_for_search(line_clean).rstrip(":")

        # Formato: "Título REFORZAMIENTO..."
        if line_norm.startswith(label_norm + " "):
            prefix_len = len(label)
            value = line_clean[prefix_len:].lstrip(" :\t")
            if value:
                return clean_multiline(value)

        # Formato: "Tipo de contenido: Evento"
        if line_norm.startswith(label_norm + ":"):
            value = line_clean.split(":", 1)[-1]
            return clean_multiline(value)

        # Formato:
        # Título
        # REFORZAMIENTO...
        if line_norm == label_norm:
            collected = []

            for j in range(i + 1, len(lines)):
                next_line = clean_text(lines[j])

                if multiline:
                    if is_known_label_line(next_line):
                        break
                    collected.append(next_line)
                else:
                    if is_known_label_line(next_line):
                        break
                    collected.append(next_line)
                    break

            return clean_multiline("\n".join(collected))

    return ""


def extract_fields_from_sharepoint_tables(soup: BeautifulSoup) -> Dict[str, str]:
    fields: Dict[str, str] = {}

    # Caso típico SharePoint: td.ms-formlabel + td.ms-formbody
    for label_cell in soup.select("td.ms-formlabel, th.ms-formlabel"):
        label = clean_text(label_cell.get_text(" ", strip=True)).replace(":", "").strip()
        if not label:
            continue

        value_cell = label_cell.find_next_sibling(["td", "th"])
        if not value_cell:
            continue

        value = clean_multiline(value_cell.get_text("\n", strip=True))
        if label and value:
            fields[label] = value

    # Fallback por filas.
    for row in soup.find_all("tr"):
        cells = row.find_all(["td", "th"], recursive=False)
        if len(cells) < 2:
            continue

        label = clean_text(cells[0].get_text(" ", strip=True)).replace(":", "").strip()
        value = clean_multiline(cells[-1].get_text("\n", strip=True))

        if label and value:
            for known in KNOWN_LABELS:
                if normalize_for_search(label) == normalize_for_search(known):
                    fields[known] = choose_longer(fields.get(known, ""), value)

    return fields


def extract_created_modified_from_text(text: str) -> Tuple[str, str, str, str]:
    creado = ""
    modificado = ""
    creado_por = ""
    modificado_por = ""

    text_clean = clean_text(text)

    m = re.search(
        r"Creado el\s+(.+?)\s+por\s+(.+?)(?=Última modificación|Ultima modificacion|$)",
        text_clean,
        flags=re.IGNORECASE,
    )
    if m:
        creado = clean_text(m.group(1))
        creado_por = clean_text(m.group(2))

    m = re.search(
        r"(?:Última modificación realizada el|Ultima modificacion realizada el)\s+(.+?)\s+por\s+(.+?)(?=Utilice esta página|Utilice esta pagina|Nombre|$)",
        text_clean,
        flags=re.IGNORECASE,
    )
    if m:
        modificado = clean_text(m.group(1))
        modificado_por = clean_text(m.group(2))

    return creado, modificado, creado_por, modificado_por


def parse_datetime_any(value) -> Tuple[str, str, str, Optional[datetime]]:
    raw = clean_text(value)

    if not raw:
        return "", "", "", None

    raw = raw.replace(".", "")
    raw = raw.replace("a. m.", "AM").replace("p. m.", "PM")
    raw = raw.replace("a.m.", "AM").replace("p.m.", "PM")
    raw = raw.replace("a m", "AM").replace("p m", "PM")
    raw = raw.replace("Z", "")
    raw = raw.replace("T", " ")

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %I:%M:%S %p",
        "%d/%m/%Y %I:%M %p",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(raw, fmt)
            return dt.strftime("%d/%m/%Y"), dt.strftime("%H:%M"), dt.isoformat(), dt
        except ValueError:
            pass

    date_match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", raw)
    time_match = re.search(r"(\d{1,2}:\d{2})", raw)

    fecha = date_match.group(1) if date_match else ""
    hora = time_match.group(1) if time_match else ""

    parsed_dt = None

    if fecha:
        try:
            parsed_date = datetime.strptime(fecha, "%d/%m/%Y").date()
            fecha = parsed_date.strftime("%d/%m/%Y")

            if hora:
                parsed_time = datetime.strptime(hora, "%H:%M").time()
                parsed_dt = datetime.combine(parsed_date, parsed_time)

        except Exception:
            pass

    return fecha, hora, parsed_dt.isoformat() if parsed_dt else "", parsed_dt


def extract_motivo(description: str, title: str = "") -> str:
    text = clean_text(description)

    patterns = [
        r"Motivo\s*:\s*(.*?)(?=Zonas afectadas|Zona afectada|Zonas\s*:|Las zonas afectadas|Subestaciones|SED|Categoría|$)",
        r"MOTIVO\s*:\s*(.*?)(?=ZONAS AFECTADAS|Zona afectada|Zonas afectadas|Zonas\s*:|Subestaciones|SED|Categoría|$)",
        r"Trabajos\s+a\s+realizar\s*:\s*(.*?)(?=Zonas afectadas|Zona afectada|Subestaciones|SED|Categoría|$)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            motivo = clean_text(match.group(1))
            if motivo:
                return motivo

    return clean_text(title)


def extract_zona_afectada(description: str, location: str = "") -> str:
    text = clean_text(description)

    patterns = [
        r"Zonas afectadas\s*:\s*(.*?)(?=Subestaciones eléctricas|Subestaciones|SED|Categoría|Tipo de contenido|$)",
        r"Zona afectada\s*:\s*(.*?)(?=Subestaciones eléctricas|Subestaciones|SED|Categoría|Tipo de contenido|$)",
        r"ZONA AFECTADA\s*:\s*(.*?)(?=Subestaciones eléctricas|Subestaciones|SED|Categoría|Tipo de contenido|$)",
        r"Zonas\s*:\s*(.*?)(?=Subestaciones eléctricas|Subestaciones|SED|Categoría|Tipo de contenido|$)",
        r"Las zonas afectadas.*?serán\s*:\s*(.*?)(?=Subestaciones eléctricas|Subestaciones|SED|Categoría|Tipo de contenido|$)",
        r"(Urbanizaciones\s+del\s+distrito\s+de\s+.*?)(?=Subestaciones eléctricas|Subestaciones|SED|Categoría|Tipo de contenido|$)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            zona = clean_text(match.group(1))
            zona = re.sub(r"\bde\s+de\b", "de", zona, flags=re.IGNORECASE)
            if zona:
                return zona

    return clean_text(location)


def extract_subestaciones(description: str) -> str:
    text = clean_text(description)

    patterns = [
        r"Subestaciones eléctricas de distribución \(SED\).*?:\s*(.*?)(?=Categoría|Tipo de contenido|$)",
        r"Subestaciones eléctricas\s*:\s*(.*?)(?=SED|Categoría|Tipo de contenido|$)",
        r"Subestaciones\s*:\s*(.*?)(?=SED|Categoría|Tipo de contenido|$)",
        r"SED\s*:\s*(.*?)(?=Categoría|Tipo de contenido|$)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            return clean_text(match.group(1))

    return ""


def extract_distritos(*parts: str) -> str:
    text = clean_text(" ".join(p for p in parts if p))
    text = re.sub(r"\bde\s+de\b", "de", text, flags=re.IGNORECASE)

    patterns = [
        r"\bDistrito\s+de\s+([^:;().,\n]+)",
        r"\bdistrito\s+de\s+([^:;().,\n]+)",
        r"\bDistrito\s+([^:;().,\n]+)",
        r"\bdistrito\s+([^:;().,\n]+)",
        r"\bDistritos\s+de\s+([^:;.\n]+)",
        r"\bdistritos\s+de\s+([^:;.\n]+)",
        r"\bUrbanizaciones\s+del\s+distrito\s+de\s+([^:;().,\n]+)",
        r"\burbanizaciones\s+del\s+distrito\s+de\s+([^:;().,\n]+)",
        r"\bZonas afectadas distrito de\s+([^:;().,\n]+)",
        r"\bzonas afectadas distrito de\s+([^:;().,\n]+)",
    ]

    found: List[str] = []

    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            raw = clean_text(match.group(1))
            raw = re.sub(r"\bAv\b.*$", "", raw, flags=re.IGNORECASE).strip()
            raw = re.sub(r"\bCalle\b.*$", "", raw, flags=re.IGNORECASE).strip()
            raw = re.sub(r"\s+y\s+", ", ", raw, flags=re.IGNORECASE)

            pieces = [p.strip(" .,-") for p in raw.split(",") if p.strip(" .,-")]

            for piece in pieces:
                if len(piece) < 2:
                    continue

                piece_norm = normalize_for_search(piece)
                existing_norms = [normalize_for_search(x) for x in found]

                if piece_norm not in existing_norms:
                    found.append(piece)

    return ", ".join(found)


def parse_display_form_html(
    item_id: int,
    html_text: str,
    fuente: str,
    rendered_text: str = "",
) -> Optional[Dict[str, str]]:
    soup = BeautifulSoup(html_text, "html.parser")
    page_text = rendered_text or get_page_text(soup)
    page_text = normalize_page_text_for_fields(page_text)

    if len(page_text) < 100:
        return None

    table_fields = extract_fields_from_sharepoint_tables(soup)

    fields: Dict[str, str] = {}

    for label in KNOWN_LABELS:
        from_table = table_fields.get(label, "")
        from_text = find_value_after_label(
            page_text,
            label,
            multiline=(label == "Descripción"),
        )

        if label == "Descripción":
            fields[label] = choose_longer(from_table, from_text)
        else:
            fields[label] = clean_text(from_table or from_text)

    creado, modificado, creado_por, modificado_por = extract_created_modified_from_text(page_text)

    if creado and not fields.get("Creado"):
        fields["Creado"] = creado
    if modificado and not fields.get("Modificado"):
        fields["Modificado"] = modificado
    if creado_por and not fields.get("Creado por"):
        fields["Creado por"] = creado_por
    if modificado_por and not fields.get("Modificado por"):
        fields["Modificado por"] = modificado_por

    title = clean_text(fields.get("Título", ""))
    location = clean_text(fields.get("Ubicación", ""))
    start_raw = clean_text(fields.get("Hora de inicio", ""))
    end_raw = clean_text(fields.get("Hora de finalización", ""))
    description = clean_multiline(fields.get("Descripción", ""))

    fecha, hora_inicio, inicio_iso, start_dt = parse_datetime_any(start_raw)
    _, hora_fin, fin_iso, _ = parse_datetime_any(end_raw)

    motivo = extract_motivo(description, title)
    zona = extract_zona_afectada(description, location)
    subestaciones = extract_subestaciones(description)
    distrito = extract_distritos(location, zona, description, title)

    if not any([fecha, hora_inicio, hora_fin, title, location, description, zona, motivo]):
        return None

    item_id_text = str(item_id)
    url = DISPFORM_URL.format(id=item_id_text)

    row_hash = content_hash(
        item_id_text,
        title,
        location,
        start_raw,
        end_raw,
        description,
    )

    return {
        "id": item_id_text,
        "url": url,
        "fecha": fecha,
        "hora_inicio": hora_inicio,
        "hora_fin": hora_fin,
        "inicio_raw": start_raw,
        "fin_raw": end_raw,
        "inicio_iso": inicio_iso,
        "fin_iso": fin_iso,
        "titulo": title,
        "ubicacion": location,
        "distrito": distrito,
        "zona_afectada": zona,
        "motivo": motivo,
        "subestaciones": subestaciones,
        "descripcion": description,
        "categoria": clean_text(fields.get("Categoría", "")),
        "todo_el_dia": clean_text(fields.get("Todo el día", "")),
        "periodicidad": clean_text(fields.get("Periodicidad", "")),
        "datos_adjuntos": clean_text(fields.get("Datos adjuntos", "")),
        "tipo_contenido": clean_text(fields.get("Tipo de contenido", "")),
        "creado": clean_text(fields.get("Creado", "")),
        "modificado": clean_text(fields.get("Modificado", "")),
        "creado_por": clean_text(fields.get("Creado por", "")),
        "modificado_por": clean_text(fields.get("Modificado por", "")),
        "fuente_extraccion": fuente,
        "hash": row_hash,
        "_dt": start_dt,
    }


def fetch_display_form_with_requests(
    session: requests.Session,
    item_id: int,
    log,
    html_dir: Optional[Path],
) -> Optional[Dict[str, str]]:
    url = DISPFORM_URL.format(id=item_id)

    try:
        response = session.get(url, headers=HEADERS_HTML, timeout=REQUEST_TIMEOUT)
        log.write(f"[DISPFORM REQUESTS] ID {item_id} -> HTTP {response.status_code}\n")

        if response.status_code != 200:
            return None

        if html_dir is not None:
            html_path = html_dir / f"{item_id}_requests.html"
            html_path.write_text(response.text, encoding="utf-8", errors="ignore")

        return parse_display_form_html(
            item_id=item_id,
            html_text=response.text,
            fuente="DISPFORM_REQUESTS",
        )

    except Exception as exc:
        log.write(f"[DISPFORM REQUESTS EXCEPTION] ID {item_id}: {repr(exc)}\n")
        return None


def fetch_display_form_with_playwright(
    page,
    item_id: int,
    log,
    html_dir: Optional[Path],
) -> Optional[Dict[str, str]]:
    """
    Abre el ID real en el navegador y extrae el HTML + texto renderizado.
    Esto es más robusto cuando requests no trae completo o la página depende de la sesión.
    """

    url = DISPFORM_URL.format(id=item_id)

    try:
        page.goto(url, wait_until="networkidle", timeout=PLAYWRIGHT_TIMEOUT)
        page.wait_for_timeout(1200)

        html_text = page.content()

        try:
            rendered_text = page.locator("body").inner_text(timeout=10000)
        except Exception:
            rendered_text = ""

        log.write(f"[DISPFORM PLAYWRIGHT] ID {item_id} OK\n")

        if html_dir is not None:
            html_path = html_dir / f"{item_id}_playwright.html"
            html_path.write_text(html_text, encoding="utf-8", errors="ignore")

            txt_path = html_dir / f"{item_id}_playwright.txt"
            txt_path.write_text(rendered_text or "", encoding="utf-8", errors="ignore")

        return parse_display_form_html(
            item_id=item_id,
            html_text=html_text,
            rendered_text=rendered_text,
            fuente="DISPFORM_PLAYWRIGHT",
        )

    except Exception as exc:
        log.write(f"[DISPFORM PLAYWRIGHT EXCEPTION] ID {item_id}: {repr(exc)}\n")
        return None


def is_event_complete_enough(event: Optional[Dict[str, str]]) -> bool:
    if not event:
        return False

    descripcion = clean_text(event.get("descripcion", ""))
    zona = clean_text(event.get("zona_afectada", ""))
    motivo = clean_text(event.get("motivo", ""))

    if not event.get("fecha") or not event.get("hora_inicio"):
        return False

    if len(descripcion) < 80:
        return False

    if len(zona) < 20:
        return False

    if len(motivo) < 5:
        return False

    return True


def merge_events(preferred: Optional[Dict[str, str]], fallback: Optional[Dict[str, str]]) -> Optional[Dict[str, str]]:
    """
    Combina dos lecturas del mismo ID.
    Se queda con el campo más largo/completo.
    """

    if not preferred:
        return fallback
    if not fallback:
        return preferred

    merged = dict(preferred)

    for key in CSV_COLUMNS:
        a = clean_multiline(preferred.get(key, ""))
        b = clean_multiline(fallback.get(key, ""))

        if key in ["descripcion", "zona_afectada", "subestaciones", "motivo"]:
            merged[key] = choose_longer(a, b)
        else:
            merged[key] = a or b

    merged["_dt"] = preferred.get("_dt") or fallback.get("_dt")
    merged["fuente_extraccion"] = (
        clean_text(preferred.get("fuente_extraccion", ""))
        + "+"
        + clean_text(fallback.get("fuente_extraccion", ""))
    ).strip("+")

    merged["hash"] = content_hash(
        merged.get("id", ""),
        merged.get("titulo", ""),
        merged.get("ubicacion", ""),
        merged.get("inicio_raw", ""),
        merged.get("fin_raw", ""),
        merged.get("descripcion", ""),
    )

    return merged


# ============================================================
# FILTROS / ORDEN
# ============================================================

def parse_row_date(row: Dict[str, str]) -> Optional[date]:
    dt_value = row.get("_dt")
    if isinstance(dt_value, datetime):
        return dt_value.date()

    fecha = row.get("fecha", "")
    if not fecha:
        return None

    try:
        return datetime.strptime(fecha, "%d/%m/%Y").date()
    except Exception:
        return None


def filter_by_date_range(
    rows: List[Dict[str, str]],
    fecha_inicio: date,
    fecha_fin: date,
) -> List[Dict[str, str]]:
    result = []

    for row in rows:
        row_date = parse_row_date(row)
        if row_date and fecha_inicio <= row_date <= fecha_fin:
            result.append(row)

    return result


def sort_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    def key(row):
        dt_value = row.get("_dt")
        if isinstance(dt_value, datetime):
            return dt_value

        try:
            return datetime.strptime(
                f"{row.get('fecha', '')} {row.get('hora_inicio', '')}",
                "%d/%m/%Y %H:%M",
            )
        except Exception:
            try:
                return datetime(1900, 1, 1) + timedelta(days=int(row.get("id", "0") or 0))
            except Exception:
                return datetime(1900, 1, 1)

    return sorted(rows, key=key)


def remove_duplicates(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    seen = set()
    unique = []

    for row in rows:
        key = row.get("hash") or content_hash(
            row.get("id", ""),
            row.get("fecha", ""),
            row.get("hora_inicio", ""),
            row.get("hora_fin", ""),
            row.get("zona_afectada", ""),
            row.get("motivo", ""),
        )

        if key in seen:
            continue

        seen.add(key)
        unique.append(row)

    return unique


def rows_for_export(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    export_rows = []
    for row in rows:
        export_rows.append({
            k: v
            for k, v in row.items()
            if not k.startswith("_")
        })
    return export_rows


# ============================================================
# GUARDADO CSV / JSON / TXT / XLSX / SQLITE
# ============================================================

def save_csv(rows: List[Dict[str, str]], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows_for_export(rows):
            writer.writerow(row)


def save_json(rows: List[Dict[str, str]], path: Path, metadata: Dict[str, str]) -> None:
    payload = {
        "metadata": metadata,
        "total": len(rows),
        "cortes": rows_for_export(rows),
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def save_txt(rows: List[Dict[str, str]], path: Path, metadata: Dict[str, str]) -> None:
    lines = []
    lines.append("EXTRACCIÓN COMPLETA DE CORTES PROGRAMADOS SEAL")
    lines.append("=" * 100)

    for key, value in metadata.items():
        lines.append(f"{key}: {value}")

    lines.append("=" * 100)
    lines.append("")

    for row in rows:
        lines.append("=" * 100)
        lines.append(f"ID: {row.get('id', '')}")
        lines.append(f"URL: {row.get('url', '')}")
        lines.append(f"Fecha: {row.get('fecha', '')}")
        lines.append(f"Hora inicio: {row.get('hora_inicio', '')}")
        lines.append(f"Hora fin: {row.get('hora_fin', '')}")
        lines.append(f"Distrito: {row.get('distrito', '')}")
        lines.append(f"Zona afectada: {row.get('zona_afectada', '')}")
        lines.append(f"Motivo: {row.get('motivo', '')}")
        lines.append(f"Ubicación: {row.get('ubicacion', '')}")
        lines.append(f"Título: {row.get('titulo', '')}")
        lines.append(f"Subestaciones: {row.get('subestaciones', '')}")
        lines.append(f"Fuente extracción: {row.get('fuente_extraccion', '')}")
        lines.append("")
        lines.append("Descripción completa:")
        lines.append(row.get("descripcion", ""))
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")


def save_xlsx(rows: List[Dict[str, str]], path: Path, metadata: Dict[str, str]) -> bool:
    """
    Guarda Excel real con tabla, filtros, congelado de cabecera y columnas ajustadas.
    Requiere openpyxl:
        pip install openpyxl
    """

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        print(f"[AVISO] No se pudo crear XLSX. Instala openpyxl. Error: {exc}")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Cortes"

    headers = [label for label, _ in XLSX_COLUMNS]
    ws.append(headers)

    for row in rows:
        ws.append([row.get(key, "") for _, key in XLSX_COLUMNS])

    # Estilo encabezados
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Formato cuerpo
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Anchos de columna
    widths = {
        "A": 8,   # ID
        "B": 12,  # Fecha
        "C": 12,  # Hora inicio
        "D": 12,  # Hora fin
        "E": 24,  # Distrito
        "F": 55,  # Zona
        "G": 45,  # Motivo
        "H": 40,  # Ubicación
        "I": 35,  # Título
        "J": 45,  # SED
        "K": 70,  # Descripción
        "L": 16,
        "M": 22,
        "N": 22,
        "O": 55,
        "P": 22,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 70

    ws.row_dimensions[1].height = 24

    # Tabla de Excel
    if ws.max_row >= 2:
        table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        tab = Table(displayName="TablaCortesSEAL", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "APP Seal - Resumen de extracción"
    ws2["A1"].font = Font(size=14, bold=True, color="1F4E78")

    r = 3
    for key, value in metadata.items():
        ws2.cell(row=r, column=1).value = key
        ws2.cell(row=r, column=2).value = value
        ws2.cell(row=r, column=1).font = Font(bold=True)
        r += 1

    ws2.cell(row=r + 1, column=1).value = "Total cortes"
    ws2.cell(row=r + 1, column=2).value = len(rows)
    ws2.cell(row=r + 1, column=1).font = Font(bold=True)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 80

    for row_cells in ws2.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)
    return True


def copy_latest(source: Path, latest: Path) -> None:
    latest.write_bytes(source.read_bytes())


def init_sqlite(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cortes (
            hash TEXT PRIMARY KEY,
            id TEXT,
            fecha TEXT,
            fecha_iso TEXT,
            hora_inicio TEXT,
            hora_fin TEXT,
            distrito TEXT,
            zona_afectada TEXT,
            motivo TEXT,
            ubicacion TEXT,
            titulo TEXT,
            descripcion TEXT,
            subestaciones TEXT,
            url TEXT,
            fuente_extraccion TEXT,
            detectado_en TEXT
        )
    """)
    conn.commit()
    return conn


def save_sqlite(rows: List[Dict[str, str]], db_path: Path, fecha_inicio: date) -> None:
    conn = init_sqlite(db_path)

    try:
        for row in rows:
            row_date = parse_row_date(row)
            fecha_iso = row_date.isoformat() if row_date else ""

            conn.execute("""
                INSERT OR REPLACE INTO cortes (
                    hash, id, fecha, fecha_iso, hora_inicio, hora_fin,
                    distrito, zona_afectada, motivo, ubicacion, titulo,
                    descripcion, subestaciones, url, fuente_extraccion, detectado_en
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                row.get("hash", ""),
                row.get("id", ""),
                row.get("fecha", ""),
                fecha_iso,
                row.get("hora_inicio", ""),
                row.get("hora_fin", ""),
                row.get("distrito", ""),
                row.get("zona_afectada", ""),
                row.get("motivo", ""),
                row.get("ubicacion", ""),
                row.get("titulo", ""),
                row.get("descripcion", ""),
                row.get("subestaciones", ""),
                row.get("url", ""),
                row.get("fuente_extraccion", ""),
                datetime.now().isoformat(timespec="seconds"),
            ))

        # Solo conserva lo activo: borra pasado anterior al mes manejado.
        conn.execute(
            "DELETE FROM cortes WHERE fecha_iso <> '' AND fecha_iso < ?",
            (fecha_inicio.isoformat(),),
        )

        conn.commit()
    finally:
        conn.close()


def cleanup_old_result_files(output_dir: Path, keep_days: int = 60) -> None:
    patterns = [
        "cortes_seal_*.csv",
        "cortes_seal_*.json",
        "cortes_seal_*.txt",
        "cortes_seal_*.xlsx",
        "extraccion_*.log",
    ]

    cutoff = datetime.now() - timedelta(days=keep_days)

    for pattern in patterns:
        for path in output_dir.glob(pattern):
            if "latest" in path.name:
                continue

            try:
                modified = datetime.fromtimestamp(path.stat().st_mtime)
                if modified < cutoff:
                    path.unlink(missing_ok=True)
            except Exception:
                pass


# ============================================================
# CONSOLA
# ============================================================

def print_table(rows: List[Dict[str, str]], max_rows: int = 80) -> None:
    if not rows:
        print("No se encontraron cortes dentro del rango de fechas.")
        return

    columns = [
        ("ID", "id", 6),
        ("Fecha", "fecha", 10),
        ("Inicio", "hora_inicio", 7),
        ("Fin", "hora_fin", 7),
        ("Distrito", "distrito", 24),
        ("Zona afectada", "zona_afectada", 58),
        ("Motivo", "motivo", 45),
    ]

    visible = rows[:max_rows]
    matrix = []

    for row in visible:
        matrix.append([
            shorten_console(row.get(key, ""), width)
            for _, key, width in columns
        ])

    headers = [c[0] for c in columns]
    widths = []

    for i, header in enumerate(headers):
        width = len(header)
        for row in matrix:
            width = max(width, len(row[i]))
        widths.append(width)

    def line(values):
        return " | ".join(
            str(values[i]).ljust(widths[i])
            for i in range(len(values))
        )

    print(line(headers))
    print("-+-".join("-" * w for w in widths))

    for row in matrix:
        print(line(row))

    if len(rows) > max_rows:
        print(
            f"\nMostrando {max_rows} de {len(rows)} registros. "
            f"Revisa el Excel/CSV para ver todo."
        )


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="APP Seal - extractor final completo y dinámico de cortes programados."
    )

    parser.add_argument(
        "--output-dir",
        default=OUTPUT_DIR_DEFAULT,
        help="Carpeta donde se guardarán los resultados.",
    )

    parser.add_argument(
        "--fecha-consulta",
        default=None,
        help="Fecha base YYYY-MM-DD. Si está vacía, usa FECHA_MANUAL_DEFAULT o fecha actual.",
    )

    parser.add_argument(
        "--desde-inicio-mes",
        action="store_true",
        help="Usa el día 1 del mes de la fecha base. Útil para pruebas.",
    )

    parser.add_argument(
        "--dias-atras",
        type=int,
        default=DIAS_ATRAS_DEFAULT,
        help="Días hacia atrás desde la fecha base. Máximo 30.",
    )

    parser.add_argument(
        "--dias-adelante",
        type=int,
        default=DIAS_ADELANTE_DEFAULT,
        help="Días hacia adelante desde la fecha base. Máximo 30.",
    )

    parser.add_argument(
        "--sin-playwright",
        action="store_true",
        help="No usar navegador. Solo intenta requests estático. No recomendado.",
    )

    parser.add_argument(
        "--ver-navegador",
        action="store_true",
        help="Muestra Chromium mientras extrae. Útil para depurar.",
    )

    parser.add_argument(
        "--no-html",
        action="store_true",
        help="No guardar HTML/TXT crudo de cada evento.",
    )

    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    html_dir = output_dir / "html_eventos"

    ensure_dir(output_dir)

    if SAVE_EVENT_HTML_DEFAULT and not args.no_html:
        ensure_dir(html_dir)
    else:
        html_dir = None

    cleanup_old_result_files(output_dir, keep_days=60)

    fecha_base = resolver_fecha_base(
        fecha_cli=args.fecha_consulta,
        usar_inicio_mes=args.desde_inicio_mes,
    )

    fecha_inicio, fecha_fin, dias_atras, dias_adelante = resolver_rango_consulta(
        fecha_base=fecha_base,
        dias_atras=args.dias_atras,
        dias_adelante=args.dias_adelante,
    )

    stamp = now_stamp()

    csv_path = output_dir / f"cortes_seal_{stamp}.csv"
    json_path = output_dir / f"cortes_seal_{stamp}.json"
    txt_path = output_dir / f"cortes_seal_{stamp}.txt"
    xlsx_path = output_dir / f"cortes_seal_{stamp}.xlsx"
    log_path = output_dir / f"extraccion_{stamp}.log"
    db_path = output_dir / "historial_cortes.sqlite"

    latest_csv = output_dir / "cortes_seal_latest.csv"
    latest_json = output_dir / "cortes_seal_latest.json"
    latest_txt = output_dir / "cortes_seal_latest.txt"
    latest_xlsx = output_dir / "cortes_seal_latest.xlsx"

    print("==============================================")
    print("APP Seal - Extractor final completo")
    print("==============================================")
    print(f"Fuente: {CORTES_PAGE}")
    print(f"Carpeta resultados: {output_dir}")
    print(f"Fecha base: {fecha_base.strftime('%d/%m/%Y')}")
    print(f"Rango manejado: {fecha_inicio.strftime('%d/%m/%Y')} a {fecha_fin.strftime('%d/%m/%Y')}")
    print(f"Días atrás: {dias_atras}")
    print(f"Días adelante: {dias_adelante}")
    print("Modo: dinámico. No se setean IDs manualmente.")
    print("Extracción de detalle: abre cada ID y lee contenido completo.")
    print("")

    session = create_session()

    metadata = {
        "fecha_ejecucion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "fuente": CORTES_PAGE,
        "fecha_base": fecha_base.isoformat(),
        "fecha_inicio": fecha_inicio.isoformat(),
        "fecha_fin": fecha_fin.isoformat(),
        "dias_atras": str(dias_atras),
        "dias_adelante": str(dias_adelante),
        "metodo_ids": "",
        "ids_detectados": "0",
        "eventos_totales_extraidos": "0",
        "eventos_guardados_en_rango": "0",
        "output_dir": str(output_dir),
    }

    ids: List[int] = []
    all_rows: List[Dict[str, str]] = []
    rows: List[Dict[str, str]] = []

    p = browser = context = page = None

    with log_path.open("w", encoding="utf-8") as log:
        log.write("APP Seal - Extracción completa final\n")
        log.write("=" * 100 + "\n")
        for key, value in metadata.items():
            log.write(f"{key}: {value}\n")
        log.write("=" * 100 + "\n\n")

        try:
            print("[INFO] Abriendo página principal para cookies/sesión...")
            response = session.get(CORTES_PAGE, headers=HEADERS_HTML, timeout=REQUEST_TIMEOUT)
            log.write(f"[INIT] {CORTES_PAGE} -> HTTP {response.status_code}\n")
        except Exception as exc:
            log.write(f"[INIT EXCEPTION] {repr(exc)}\n")

        # Intento ligero
        try:
            print("[INFO] Buscando IDs dinámicos con requests...")
            ids = discover_ids_static_requests(
                session=session,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                log=log,
            )

            if ids:
                metadata["metodo_ids"] = "REQUESTS_STATIC"
                print(f"[OK] IDs encontrados con requests: {len(ids)}")
            else:
                print("[INFO] Requests no encontró IDs. Se usará Playwright.")

        except Exception as exc:
            log.write(f"[STATIC GENERAL ERROR] {repr(exc)}\n")
            print(f"[AVISO] Requests estático falló: {exc}")

        # Playwright para IDs y detalles
        if not args.sin_playwright:
            try:
                print("[INFO] Iniciando Playwright...")
                p, browser, context, page = create_playwright_context(headless=not args.ver_navegador)

                if not ids:
                    print("[INFO] Buscando IDs dinámicos renderizando el calendario...")
                    ids = discover_ids_rendered_with_playwright(
                        page=page,
                        session=session,
                        fecha_inicio=fecha_inicio,
                        fecha_fin=fecha_fin,
                        log=log,
                    )

                    if ids:
                        metadata["metodo_ids"] = "PLAYWRIGHT_RENDERED"
                        print(f"[OK] IDs encontrados con Playwright: {len(ids)}")
                    else:
                        print("[INFO] Playwright no encontró IDs.")

                sync_playwright_cookies_to_requests(context, session)

            except Exception as exc:
                log.write(f"[PLAYWRIGHT INIT/DISCOVER ERROR] {repr(exc)}\n")
                print(f"[AVISO] Playwright falló: {exc}")

        ids = sorted(set(ids))
        metadata["ids_detectados"] = str(len(ids))

        print("")
        print(f"[INFO] Total IDs dinámicos detectados: {len(ids)}")

        if ids:
            preview = ids[:30]
            suffix = "..." if len(ids) > 30 else ""
            print(f"[INFO] Preview IDs: {preview}{suffix}")
        print("")

        for index, item_id in enumerate(ids, start=1):
            # Primero intenta con requests.
            event_requests = fetch_display_form_with_requests(
                session=session,
                item_id=item_id,
                log=log,
                html_dir=html_dir,
            )

            event_final = event_requests

            # Si no está completo, refuerza con Playwright abriendo el ID real.
            if not is_event_complete_enough(event_final) and page is not None:
                event_pw = fetch_display_form_with_playwright(
                    page=page,
                    item_id=item_id,
                    log=log,
                    html_dir=html_dir,
                )
                event_final = merge_events(event_pw, event_requests)

            if event_final:
                all_rows.append(event_final)

                row_date = parse_row_date(event_final)
                in_range = row_date is not None and fecha_inicio <= row_date <= fecha_fin
                status = "EN RANGO" if in_range else "FUERA"

                zona_len = len(clean_text(event_final.get("zona_afectada", "")))
                desc_len = len(clean_text(event_final.get("descripcion", "")))

                print(
                    f"[OK-{status}] {index}/{len(ids)} "
                    f"ID {item_id} "
                    f"{event_final.get('fecha', '')} "
                    f"{event_final.get('hora_inicio', '')}-{event_final.get('hora_fin', '')} "
                    f"{shorten_console(event_final.get('distrito', ''), 28)} "
                    f"(zona {zona_len} chars, desc {desc_len} chars)"
                )
            else:
                print(f"[SKIP] {index}/{len(ids)} ID {item_id} sin datos útiles")

            time.sleep(SLEEP_BETWEEN_EVENTS)

        if browser is not None:
            try:
                browser.close()
            except Exception:
                pass
        if p is not None:
            try:
                p.stop()
            except Exception:
                pass

        all_rows = remove_duplicates(all_rows)
        all_rows = sort_rows(all_rows)

        rows = filter_by_date_range(all_rows, fecha_inicio, fecha_fin)
        rows = remove_duplicates(rows)
        rows = sort_rows(rows)

        metadata["eventos_totales_extraidos"] = str(len(all_rows))
        metadata["eventos_guardados_en_rango"] = str(len(rows))

        log.write("\nResumen final\n")
        log.write(f"Metodo IDs: {metadata.get('metodo_ids', '')}\n")
        log.write(f"IDs detectados: {len(ids)}\n")
        log.write(f"Eventos totales extraídos: {len(all_rows)}\n")
        log.write(f"Eventos en rango: {len(rows)}\n")
        log.write(f"Fin: {datetime.now().isoformat()}\n")

    save_csv(rows, csv_path)
    save_json(rows, json_path, metadata)
    save_txt(rows, txt_path, metadata)
    save_sqlite(rows, db_path, fecha_inicio=fecha_inicio)

    xlsx_ok = save_xlsx(rows, xlsx_path, metadata)

    copy_latest(csv_path, latest_csv)
    copy_latest(json_path, latest_json)
    copy_latest(txt_path, latest_txt)

    if xlsx_ok:
        copy_latest(xlsx_path, latest_xlsx)

    print("")
    print("==============================================")
    print("RESUMEN")
    print("==============================================")
    print(f"Método IDs: {metadata.get('metodo_ids') or 'NINGUNO'}")
    print(f"IDs dinámicos detectados: {len(ids)}")
    print(f"Eventos totales extraídos: {len(all_rows)}")
    print(f"Eventos guardados en rango: {len(rows)}")
    print(f"Rango guardado: {fecha_inicio.strftime('%d/%m/%Y')} a {fecha_fin.strftime('%d/%m/%Y')}")
    print("")

    print_table(rows)

    print("")
    print("==============================================")
    print("ARCHIVOS GENERADOS")
    print("==============================================")
    print(f"Excel tabla:  {xlsx_path if xlsx_ok else 'No generado. Instala openpyxl.'}")
    print(f"CSV:          {csv_path}")
    print(f"JSON:         {json_path}")
    print(f"TXT:          {txt_path}")
    print(f"SQLite:       {db_path}")
    print(f"LOG:          {log_path}")
    print(f"Excel latest: {latest_xlsx if xlsx_ok else 'No generado'}")
    print(f"CSV latest:   {latest_csv}")

    if not ids:
        print("")
        print("No se detectaron IDs dinámicos.")
        print("Instala dependencias:")
        print("pip install requests beautifulsoup4 playwright openpyxl")
        print("python -m playwright install chromium")
        print("")
        print("Depura viendo navegador:")
        print("python seal_cortes_final_v2.py --ver-navegador")

    if not rows and ids:
        print("")
        print("Se detectaron IDs, pero no quedaron cortes dentro del rango.")
        print("Para probar con junio usa:")
        print("python seal_cortes_final_v2.py --fecha-consulta 2026-06-01")
        print("")
        print("O desde el inicio del mes:")
        print("python seal_cortes_final_v2.py --desde-inicio-mes")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nProceso cancelado por el usuario.")
        sys.exit(130)
    except Exception as exc:
        print(f"\nERROR GENERAL: {repr(exc)}")
        sys.exit(1)
